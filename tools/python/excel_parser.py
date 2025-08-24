"""
Excel to MDN format converter.
Converts Excel (.xlsx) files to MDN (Markdown Numbers) format.
"""

import openpyxl
import yaml
import csv
import io
from typing import Dict, List, Any, Tuple
from utils import format_timestamp, column_index_to_letter


class ExcelToMDNConverter:
    """Convert Excel workbooks to MDN format."""
    
    def __init__(self):
        self.workbook = None
        self.mdn_content = []
    
    def convert_file(self, excel_file_path: str) -> str:
        """
        Convert Excel file to MDN format.
        
        Args:
            excel_file_path: Path to the Excel file
            
        Returns:
            MDN format content as string
        """
        print(f"Converting {excel_file_path} to MDN format...")
        
        # Load Excel workbook
        self.workbook = openpyxl.load_workbook(excel_file_path, data_only=False)
        
        # Generate MDN content
        self._generate_header(excel_file_path)
        self._generate_sheets()
        self._generate_formulas()
        self._generate_formatting()
        self._add_end_marker()
        
        # Join all sections
        mdn_content = '\n'.join(self.mdn_content)
        
        print(f"✓ Successfully converted to MDN format")
        print(f"✓ Generated {len(self.workbook.sheetnames)} sheets")
        
        return mdn_content
    
    def _generate_header(self, source_file: str):
        """Generate YAML header section."""
        print("  Generating YAML header...")
        
        # Extract filename without path
        import os
        filename = os.path.basename(source_file)
        
        header_data = {
            'source': filename,
            'version': '1.0',
            'created': format_timestamp(),
            'sheets': self.workbook.sheetnames
        }
        
        # Add optional context section
        context_data = {
            'purpose': 'excel_to_mdn_conversion',
            'keyMetrics': ['data_integrity', 'formula_preservation'],
            'businessRules': [
                'All formulas must be preserved',
                'Data types maintained through formatting',
                'Sheet structure preserved'
            ]
        }
        
        # Generate header section
        self.mdn_content.append('--- MDN:HEADER YAML')
        self.mdn_content.append(yaml.dump(header_data, default_flow_style=False, sort_keys=False))
        self.mdn_content.append('---')
        
        # Add context section
        self.mdn_content.append('# optional context section')
        self.mdn_content.append(yaml.dump(context_data, default_flow_style=False, sort_keys=False))
        self.mdn_content.append('---')
        
        print(f"    ✓ Header generated with {len(self.workbook.sheetnames)} sheets")
    
    def _generate_sheets(self):
        """Generate CSV sheet sections."""
        print("  Generating CSV sheet sections...")
        
        for sheet_name in self.workbook.sheetnames:
            print(f"    Processing sheet: {sheet_name}")
            sheet = self.workbook[sheet_name]
            
            # Convert sheet to CSV
            csv_data = self._sheet_to_csv(sheet)
            
            # Add sheet section
            self.mdn_content.append(f'--- MDN:SHEET CSV name={sheet_name}')
            self.mdn_content.append(csv_data)
            self.mdn_content.append('---')
            
            print(f"      ✓ Sheet '{sheet_name}' converted to CSV")
    
    def _sheet_to_csv(self, sheet) -> str:
        """
        Convert Excel sheet to CSV format.
        
        Args:
            sheet: OpenPyXL worksheet object
            
        Returns:
            CSV data as string
        """
        # Get the used range
        max_row = sheet.max_row
        max_col = sheet.max_column
        
        if max_row == 0 or max_col == 0:
            return ""  # Empty sheet
        
        # Convert to CSV
        csv_buffer = io.StringIO()
        csv_writer = csv.writer(csv_buffer)
        
        for row in range(1, max_row + 1):
            row_data = []
            for col in range(1, max_col + 1):
                cell = sheet.cell(row=row, column=col)
                value = cell.value
                
                # Handle different data types
                if value is None:
                    row_data.append("")
                elif isinstance(value, (int, float)):
                    row_data.append(str(value))
                else:
                    row_data.append(str(value))
            
            csv_writer.writerow(row_data)
        
        return csv_buffer.getvalue().strip()
    
    def _generate_formulas(self):
        """Generate formulas JSON section."""
        print("  Generating formulas section...")
        
        formulas = {}
        total_formulas = 0
        
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            
            for row in range(1, sheet.max_row + 1):
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row, column=col)
                    
                    if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith('='):
                        # This is a formula
                        cell_ref = f"{column_index_to_letter(col)}{row}"
                        formula_key = f"{sheet_name}!{cell_ref}"
                        formulas[formula_key] = cell.value
                        total_formulas += 1
        
        # Add formulas section
        self.mdn_content.append('--- MDN:FORMULAS JSON')
        self.mdn_content.append(yaml.dump(formulas, default_flow_style=False, sort_keys=False))
        self.mdn_content.append('---')
        
        print(f"    ✓ Formulas section generated with {total_formulas} formulas")
    
    def _generate_formatting(self):
        """Generate formatting JSON section."""
        print("  Generating formatting section...")
        
        formatting = {}
        total_formatting_rules = 0
        
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            
            for row in range(1, sheet.max_row + 1):
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row, column=col)
                    cell_ref = f"{column_index_to_letter(col)}{row}"
                    
                    # Check for number format
                    if cell.number_format and cell.number_format != 'General':
                        format_key = f"{sheet_name}!{cell_ref}"
                        if format_key not in formatting:
                            formatting[format_key] = {}
                        formatting[format_key]['numberFormat'] = cell.number_format
                        total_formatting_rules += 1
                    
                    # Check for font properties
                    if cell.font:
                        format_key = f"{sheet_name}!{cell_ref}"
                        if format_key not in formatting:
                            formatting[format_key] = {}
                        
                        if cell.font.bold:
                            formatting[format_key]['bold'] = True
                            total_formatting_rules += 1
                        
                        if cell.font.italic:
                            formatting[format_key]['italic'] = True
                            total_formatting_rules += 1
                        
                        # Only include color if it's not the default theme color (theme=1 is typically black)
                        if (cell.font.color and 
                            hasattr(cell.font.color, 'theme') and 
                            cell.font.color.theme != 1):
                            formatting[format_key]['theme'] = cell.font.color.theme
                            total_formatting_rules += 1
                        
                        # Handle RGB colors if present
                        if (cell.font.color and 
                            hasattr(cell.font.color, 'rgb') and 
                            cell.font.color.rgb and
                            hasattr(cell.font.color, 'type') and
                            cell.font.color.type == 'rgb'):
                            try:
                                rgb = str(cell.font.color.rgb)
                                if rgb.startswith('FF'):  # Remove alpha channel
                                    rgb = rgb[2:]
                                formatting[format_key]['color'] = f"#{rgb}"
                                total_formatting_rules += 1
                            except Exception:
                                # Skip color if there's an issue
                                pass
        
        # Add formatting section (only if there are formatting rules)
        if formatting:
            # Remove empty formatting objects
            clean_formatting = {k: v for k, v in formatting.items() if v}
            
            if clean_formatting:
                self.mdn_content.append('--- MDN:FORMAT JSON')
                self.mdn_content.append(yaml.dump(clean_formatting, default_flow_style=False, sort_keys=False))
                self.mdn_content.append('---')
                print(f"    ✓ Formatting section generated with {len(clean_formatting)} rules")
            else:
                print("    ✓ No meaningful formatting rules found, skipping formatting section")
        else:
            print("    ✓ No formatting rules found, skipping formatting section")
    
    def _add_end_marker(self):
        """Add END DOCUMENT marker."""
        self.mdn_content.append('END DOCUMENT')


def excel_to_mdn(excel_file_path: str) -> str:
    """
    Convenience function to convert Excel file to MDN format.
    
    Args:
        excel_file_path: Path to the Excel file
        
    Returns:
        MDN format content as string
    """
    converter = ExcelToMDNConverter()
    return converter.convert_file(excel_file_path)


if __name__ == "__main__":
    # Example usage
    import sys
    
    if len(sys.argv) != 2:
        print("Usage: python excel_parser.py <excel_file>")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    try:
        mdn_content = excel_to_mdn(excel_file)
        
        # Save to file
        output_file = excel_file.replace('.xlsx', '.mdn')
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(mdn_content)
        
        print(f"\nMDN file saved as: {output_file}")
        
    except Exception as e:
        print(f"Error converting file: {e}")
        sys.exit(1)
