"""
MDN to Excel format converter.
Converts MDN (Markdown Numbers) format to Excel (.xlsx) files.
"""

import openpyxl
import yaml
import csv
import io
import re
from typing import Dict, List, Any, Tuple
from utils import parse_cell_reference, column_letter_to_index


class MDNToExcelConverter:
    """Convert MDN format to Excel workbooks."""
    
    def __init__(self):
        self.workbook = None
        self.sheets_data = {}
        self.formulas = {}
        self.formatting = {}
        self.header_data = {}
    
    def convert_content(self, mdn_content: str, output_file_path: str):
        """
        Convert MDN content to Excel file.
        
        Args:
            mdn_content: MDN format content as string
            output_file_path: Path for the output Excel file
        """
        print(f"Converting MDN content to Excel format...")
        
        # Parse MDN content
        self._parse_mdn_content(mdn_content)
        
        # Create Excel workbook
        self._create_workbook()
        
        # Apply data to sheets
        self._apply_sheet_data()
        
        # Apply formulas
        self._apply_formulas()
        
        # Apply formatting
        self._apply_formatting()
        
        # Save workbook
        self.workbook.save(output_file_path)
        
        print(f"✓ Successfully converted to Excel format")
        print(f"✓ Saved as: {output_file_path}")
        print(f"✓ Created {len(self.workbook.sheetnames)} sheets")
    
    def _parse_mdn_content(self, content: str):
        """Parse MDN content into sections."""
        print("  Parsing MDN content...")
        
        # Split content into sections
        sections = content.split('---')
        
        for section in sections:
            section = section.strip()
            if not section:
                continue
            
            if section.startswith('MDN:HEADER YAML'):
                self._parse_header_section(section)
            elif section.startswith('MDN:SHEET CSV'):
                self._parse_sheet_section(section)
            elif section.startswith('MDN:FORMULAS JSON'):
                self._parse_formulas_section(section)
            elif section.startswith('MDN:FORMAT JSON'):
                self._parse_formatting_section(section)
        
        print(f"    ✓ Parsed {len(self.sheets_data)} sheets")
        print(f"    ✓ Found {len(self.formulas)} formulas")
        print(f"    ✓ Found {len(self.formatting)} formatting rules")
    
    def _parse_header_section(self, section: str):
        """Parse YAML header section."""
        # Extract YAML content after the section marker
        yaml_content = section.replace('MDN:HEADER YAML', '').strip()
        
        # Remove context section if present
        if '# optional context section' in yaml_content:
            yaml_content = yaml_content.split('# optional context section')[0].strip()
        
        try:
            self.header_data = yaml.safe_load(yaml_content)
            print(f"      ✓ Header parsed: {self.header_data.get('source', 'Unknown')}")
        except yaml.YAMLError as e:
            print(f"      ⚠ Warning: Could not parse header YAML: {e}")
            self.header_data = {'source': 'unknown.xlsx', 'sheets': []}
    
    def _parse_sheet_section(self, section: str):
        """Parse CSV sheet section."""
        # Extract sheet name and CSV data
        lines = section.split('\n')
        sheet_name = None
        csv_data = []
        
        for line in lines:
            if line.startswith('MDN:SHEET CSV'):
                # Extract sheet name from "name=SheetName"
                match = re.search(r'name=([^\s]+)', line)
                if match:
                    sheet_name = match.group(1)
            elif line.strip() and not line.startswith('MDN:'):
                csv_data.append(line)
        
        if sheet_name and csv_data:
            self.sheets_data[sheet_name] = csv_data
            print(f"      ✓ Sheet '{sheet_name}' parsed with {len(csv_data)} rows")
    
    def _parse_formulas_section(self, section: str):
        """Parse formulas JSON section."""
        # Extract JSON content after the section marker
        json_content = section.replace('MDN:FORMULAS JSON', '').strip()
        
        try:
            self.formulas = yaml.safe_load(json_content) or {}
            print(f"      ✓ Formulas parsed: {len(self.formulas)} formulas")
        except yaml.YAMLError as e:
            print(f"      ⚠ Warning: Could not parse formulas JSON: {e}")
            self.formulas = {}
    
    def _parse_formatting_section(self, section: str):
        """Parse formatting JSON section."""
        # Extract JSON content after the section marker
        json_content = section.replace('MDN:FORMAT JSON', '').strip()
        
        try:
            self.formatting = yaml.safe_load(json_content) or {}
            print(f"      ✓ Formatting parsed: {len(self.formatting)} rules")
        except yaml.YAMLError as e:
            print(f"      ⚠ Warning: Could not parse formatting JSON: {e}")
            self.formatting = {}
    
    def _create_workbook(self):
        """Create Excel workbook with sheets."""
        print("  Creating Excel workbook...")
        
        self.workbook = openpyxl.Workbook()
        
        # Remove default sheet
        default_sheet = self.workbook.active
        self.workbook.remove(default_sheet)
        
        # Create sheets based on parsed data
        for sheet_name in self.sheets_data.keys():
            self.workbook.create_sheet(title=sheet_name)
        
        print(f"    ✓ Created workbook with {len(self.workbook.sheetnames)} sheets")
    
    def _apply_sheet_data(self):
        """Apply CSV data to Excel sheets."""
        print("  Applying sheet data...")
        
        for sheet_name, csv_data in self.sheets_data.items():
            sheet = self.workbook[sheet_name]
            
            # Parse CSV data and apply to sheet
            for row_idx, csv_row in enumerate(csv_data, 1):
                # Parse CSV row (handle quoted values)
                row_values = list(csv.reader([csv_row]))[0]
                
                for col_idx, value in enumerate(row_values, 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    
                    # Convert value to appropriate type
                    if value == "":
                        cell.value = None
                    elif value.isdigit():
                        cell.value = int(value)
                    elif value.replace('.', '').replace('-', '').isdigit():
                        cell.value = float(value)
                    else:
                        cell.value = value
            
            print(f"    ✓ Applied data to sheet '{sheet_name}'")
    
    def _apply_formulas(self):
        """Apply formulas to Excel sheets."""
        print("  Applying formulas...")
        
        for formula_key, formula_value in self.formulas.items():
            try:
                # Parse formula key (e.g., "Sheet1!A1")
                if '!' in formula_key:
                    sheet_name, cell_ref = formula_key.split('!', 1)
                    
                    if sheet_name in self.workbook.sheetnames:
                        sheet = self.workbook[sheet_name]
                        
                        # Parse cell reference
                        col_letter, row_num = self._parse_cell_ref(cell_ref)
                        col_idx = column_letter_to_index(col_letter)
                        
                        # Apply formula
                        cell = sheet.cell(row=row_num, column=col_idx)
                        cell.value = formula_value
                        
            except Exception as e:
                print(f"      ⚠ Warning: Could not apply formula {formula_key}: {e}")
        
        print(f"    ✓ Applied {len(self.formulas)} formulas")
    
    def _parse_cell_ref(self, cell_ref: str) -> Tuple[str, int]:
        """Parse cell reference like 'A1' into column letter and row number."""
        match = re.match(r'([A-Z]+)(\d+)', cell_ref)
        if match:
            return match.group(1), int(match.group(2))
        else:
            raise ValueError(f"Invalid cell reference: {cell_ref}")
    
    def _apply_formatting(self):
        """Apply formatting to Excel sheets."""
        print("  Applying formatting...")
        
        for format_key, format_rules in self.formatting.items():
            try:
                # Parse format key (e.g., "Sheet1!A1")
                if '!' in format_key:
                    sheet_name, cell_ref = format_key.split('!', 1)
                    
                    if sheet_name in self.workbook.sheetnames:
                        sheet = self.workbook[sheet_name]
                        
                        # Parse cell reference
                        col_letter, row_num = self._parse_cell_ref(cell_ref)
                        col_idx = column_letter_to_index(col_letter)
                        
                        # Apply formatting
                        cell = sheet.cell(row=row_num, column=col_idx)
                        
                        # Apply number format
                        if 'numberFormat' in format_rules:
                            cell.number_format = format_rules['numberFormat']
                        
                        # Apply font properties
                        if 'bold' in format_rules and format_rules['bold']:
                            cell.font = openpyxl.styles.Font(bold=True)
                        
                        if 'italic' in format_rules and format_rules['italic']:
                            if cell.font:
                                cell.font.italic = True
                            else:
                                cell.font = openpyxl.styles.Font(italic=True)
                        
                        if 'color' in format_rules:
                            color = format_rules['color']
                            if color.startswith('#'):
                                color = 'FF' + color[1:]  # Add alpha channel
                            if cell.font:
                                cell.font.color = openpyxl.styles.Color(rgb=color)
                            else:
                                cell.font = openpyxl.styles.Color(rgb=color)
                        
            except Exception as e:
                print(f"      ⚠ Warning: Could not apply formatting {format_key}: {e}")
        
        print(f"    ✓ Applied {len(self.formatting)} formatting rules")


def mdn_to_excel(mdn_content: str, output_file_path: str):
    """
    Convenience function to convert MDN content to Excel file.
    
    Args:
        mdn_content: MDN format content as string
        output_file_path: Path for the output Excel file
    """
    converter = MDNToExcelConverter()
    converter.convert_content(mdn_content, output_file_path)


if __name__ == "__main__":
    # Example usage
    import sys
    
    if len(sys.argv) != 3:
        print("Usage: python mdn_parser.py <mdn_file> <output_excel_file>")
        sys.exit(1)
    
    mdn_file = sys.argv[1]
    output_excel = sys.argv[2]
    
    try:
        # Read MDN file
        with open(mdn_file, 'r', encoding='utf-8') as f:
            mdn_content = f.read()
        
        # Convert to Excel
        mdn_to_excel(mdn_content, output_excel)
        
        print(f"\nExcel file saved as: {output_excel}")
        
    except Exception as e:
        print(f"Error converting file: {e}")
        sys.exit(1)
